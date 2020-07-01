# -*- coding: utf-8 -*-
"""
Created on Tue Jun 23 15:46:43 2020

@author: 6989
"""

def RateInput(ws, excelRow, data, i, Cover):
    if Cover:
        ws.cell(row = excelRow, column = 1).value = str(data["資料日期"][i])
    ws.cell(row = excelRow, column = 4).value = float(data["銀行買入"][i])
    ws.cell(row = excelRow, column = 5).value = float(data["銀行賣出"][i])    

import datetime
year = datetime.datetime.today().year
month = datetime.datetime.today().strftime("%m")
day = datetime.datetime.today().strftime("%d")
hour = datetime.datetime.today().hour
minute = datetime.datetime.today().minute
date = f'{year}_{month}_{day}'
today = f'{year}/{month}/{day}'

import urllib.request as req
import bs4
import pandas as pd
import csv
import openpyxl as xl

    
pageURL = f"https://rate.bot.com.tw/gold/chart/year/TWD"
    
# 建立一個 request 物件，附加 request headers 的資訊
request = req.Request(pageURL, headers={        
        "user-agent":"mozilla/5.0 (windows nt 10.0; win64; x64) applewebkit/537.36 (khtml, like gecko) chrome/80.0.3987.163 safari/537.36"
        })
    
with req.urlopen(request) as response:
    data =response.read().decode("utf-8")
# print(data)
    
root = bs4.BeautifulSoup(data, "html.parser")
    
downloadCSV = root.find("a", string = "下載 Excel (CSV) 檔")
csvLink = downloadCSV["href"]
    
#下載檔案
fileName = f'Gold_{date}'
req.urlretrieve("https://rate.bot.com.tw"+csvLink, f"./Gold/{fileName}.csv")
    
    
    
with open(f"./Gold/{fileName}.csv", newline='', encoding = 'utf-8') as csvfile:    
    gold = csv.reader(csvfile)        
    inputDate = []
    bankBuy = []        
    bankSold = []        
    check = False
    for row in gold:
        if(check):
            inputDate.append(row[0][:4] + "/" + row[0][4:6] + "/" + row[0][6:])
            bankBuy.append(row[3])                
            bankSold.append(row[4])                
        else:
            check = True      
    inputDate.reverse()
    bankBuy.reverse()        
    bankSold.reverse()        
        
    dic = {"資料日期":inputDate, 
           "銀行買入":bankBuy,         
           "銀行賣出":bankSold,                
           }
    data = pd.DataFrame(dic)
# print(len(data))
    
    
wb = xl.load_workbook("./Gold/Gold.xlsx")
ws = wb["Gold"]
            
    
    
excelRow = 1
    
existDate = []
while (ws.cell(row = excelRow + 1, column = 1).value != None):
    excelRow += 1
    existDate.append(ws.cell(row = excelRow, column = 1).value)
# print(existDate)
    
for i in range(len(data)):
    if data["資料日期"][i] not in existDate:
        excelRow += 1
        RateInput(ws, excelRow, data, i, True)        
    elif (today == data["資料日期"][i]):
        RateInput(ws, excelRow, data, i, True) 
        RateInput(ws, excelRow - 1, data, i - 1, False)
        RateInput(ws, excelRow - 2, data, i - 2, False)
        RateInput(ws, excelRow - 3, data, i - 3, False)
wb.save(f'./Gold/Gold.xlsx')
wb.close()

# =============================================================================
# if (today == data["資料日期"][len(data)-1]):
#     print("True")
# =============================================================================











