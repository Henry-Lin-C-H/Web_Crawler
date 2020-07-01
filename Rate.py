# -*- coding: utf-8 -*-
"""
Created on Tue Jun 23 15:46:43 2020

@author: 6989
"""

def RateInput(ws, excelRow, data, i, Cover):
    if Cover:
        ws.cell(row = excelRow, column = 1).value = str(data["資料日期"][i])
    ws.cell(row = excelRow, column = 4).value = float(data["現金買入"][i])
    ws.cell(row = excelRow, column = 5).value = float(data["即期買入"][i])
    ws.cell(row = excelRow, column = 14).value = float(data["現金賣出"][i])
    ws.cell(row = excelRow, column = 15).value = float(data["即期賣出"][i])

import datetime
year = datetime.datetime.today().year
month = datetime.datetime.today().strftime("%m")
day = datetime.datetime.today().strftime("%d")
hour = datetime.datetime.today().hour
minute = datetime.datetime.today().minute
date = f'{year}_{month}_{day}'
today = f'{year}/{month}/{day}'

import urllib.request as req
import bs4
import pandas as pd
import csv
import openpyxl as xl

RateName = ["USD", "JPY"]
for allNo in range(len(RateName)):
    # print(allNo)    
    
    pageURL = f"https://rate.bot.com.tw/xrt/quote/l6m/{RateName[allNo]}"
    
    # 建立一個 request 物件，附加 request headers 的資訊
    request = req.Request(pageURL, headers={        
            "user-agent":"mozilla/5.0 (windows nt 10.0; win64; x64) applewebkit/537.36 (khtml, like gecko) chrome/80.0.3987.163 safari/537.36"
            })
    
    with req.urlopen(request) as response:
        data =response.read().decode("utf-8")
    # print(data)
    
    root = bs4.BeautifulSoup(data, "html.parser")
    
    downloadCSV = root.find("a", string = "下載 Excel (CSV) 檔")
    csvLink = downloadCSV["href"]
    
    #下載檔案
    fileName = f'{RateName[allNo]}_{date}'
    req.urlretrieve("https://rate.bot.com.tw"+csvLink, f"./Rate/{fileName}.csv")
    
    
    
    with open(f"./Rate/{fileName}.csv", newline='', encoding = 'utf-8') as csvfile:
    # =============================================================================
    # # 用 DictReader 無法讀取 "資料日期" ，原因未知，且相同名稱的項目亦無法重複讀取
    #     rate = csv.DictReader(csvfile)
    #     dic = pd.DataFrame(rate)
    #     for row in rate:
    #         print(row["幣別"])
    # print(dic["現金"])
    # =============================================================================
    
        rate = csv.reader(csvfile)        
        inputDate = []
        MoneyBuy = []
        RateBuy = []
        MoneySold = []
        RateSold = []
        check = False
        for row in rate:
            if(check):
                inputDate.append(row[0][:4] + "/" + row[0][4:6] + "/" + row[0][6:])
                MoneyBuy.append(row[3])
                RateBuy.append(row[4])
                MoneySold.append(row[13])
                RateSold.append(row[14])
            else:
                check = True      
        inputDate.reverse()
        MoneyBuy.reverse()
        RateBuy.reverse()
        MoneySold.reverse()
        RateSold.reverse()
        
        dic = {"資料日期":inputDate, 
               "現金買入":MoneyBuy, 
               "即期買入":RateBuy, 
               "現金賣出":MoneySold, 
               "即期賣出":RateSold
               }
        data = pd.DataFrame(dic)
    # print(len(data))
    
    
    wb = xl.load_workbook("./Rate/Rate.xlsx")
    ws = wb[RateName[allNo]]
            
    
    
    excelRow = 1
    
    existDate = []
    while (ws.cell(row = excelRow + 1, column = 1).value != None):
        excelRow += 1
        existDate.append(ws.cell(row = excelRow, column = 1).value)
    # print(existDate)
    
    for i in range(len(data)):
        if data["資料日期"][i] not in existDate:
            excelRow += 1
            RateInput(ws, excelRow, data, i, True)        
        elif (today == data["資料日期"][i]):
            RateInput(ws, excelRow, data, i, True) 
            RateInput(ws, excelRow - 1, data, i - 1, False)
            RateInput(ws, excelRow - 2, data, i - 2, False)
            RateInput(ws, excelRow - 3, data, i - 3, False)
    wb.save(f'./Rate/Rate.xlsx')
    wb.close()

# =============================================================================
# if (today == data["資料日期"][len(data)-1]):
#     print("True")
# =============================================================================











