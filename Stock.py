# -*- coding: utf-8 -*-
"""
Created on Thu Jun 18 15:46:44 2020

@author: 6989
"""

import datetime
year = datetime.datetime.today().year
month = datetime.datetime.today().strftime("%m")
day = datetime.datetime.today().strftime("%d")
hour = datetime.datetime.today().hour
# minute = datetime.datetime.today().minute
minute = datetime.datetime.today().strftime("%M")
date = f'{year}_{month}_{day}'
nowTime = f'{hour}:{minute}'

stockNo = []
stockName = []
import csv
with open("選股標的.csv", newline='') as csvfile:
    designedStock = csv.DictReader(csvfile)
    for row in designedStock:
        stockNo.append(row["編號"])
        stockName.append(row["名稱"])
#        print(row["編號"], row["名稱"])

stockValue = []

import urllib.request as req
import bs4
import openpyxl as xl
import pandas as pd
for no in range(len(stockNo)):
    #抓取股票資料 (HTML)    
#    url = f'https://tw.stock.yahoo.com/q/ts?s={stockNo[no]}'
    url = f'https://tw.stock.yahoo.com/q/ts?s={stockNo[no]}&t=50'
    
    # 建立一個 Request 物件，附加 Request Headers 的資訊
    request = req.Request(url, headers={
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36"
            })
    
    with req.urlopen(request) as response:
        data = response.read()
    #    data = response.read().decode("utf-8")
    #print(data)
    
    # 解析原始碼，取得每篇文章的標題
    # 要先安裝 BeautifulSoup4   打開Python或 VisualStudioCode 寫入 pip install beautifulsoup4
    
    root = bs4.BeautifulSoup(data, "html.parser")
    #print(root.title.string)
    
    # =============================================================================
    # tds = root.find_all("td", class_="high")
    # #tds = root.find_all("td")
    # for td in tds:
    #     print(td.string)
    # =============================================================================
    
    trs = root.find_all("tr", align="center", bgcolor="#ffffff", height="25")
    
    time = []
    quantity = []  
    value = []
    variation = []   
    first = True    
    for td in trs:
        if first is True:
            stockValue.append(td.find_all("td")[4].string)
            first = False
        time.append(td.find_all("td")[0].string)
        quantity.append(td.find_all("td")[6].string)
        value.append(td.find_all("td")[4].string)
        variation.append(td.find_all("td")[5].string)
    
    
    data = pd.DataFrame({
            "Time":time,
            "Quantity":quantity,
            "Value":value,
            "Variation":variation
            }) #股票爬蟲抓下來的時間與量
#    print(data)
           
    inputHour = ["09", "10", "11", "12", "13"]
    inputMinu = ["00", "05", "10", "15", "20", "25", "30", "35", "40", "45", "50", "55"]
    allTime = [] #全部的時間
    for i in inputHour:
        for j in inputMinu:
            if i == "13" and j == "35":
                break
            elif i == "09" and j == "00":
                continue                
            else:
                allTime.append(str(i)+":"+str(j))
    #print(time)
    
    check = []
    inputTime = [] # 輸入時間
    for i in data["Time"]:
        if i in allTime:
            check.append(True)
            inputTime.append(i + ":00")
        else:
            check.append(False)
    
    del inputTime[-1]
    
    data["Calculate"] = check #確認需要計算的區間
#    print(data)
    
    calQuan = 0 # 計算量
    firstBool = True # 第一個量判斷用
    inputQuan = [] # 輸入的量
    inputValue = []
    inputVariation = []      
        
    for i in data.index:
        if check[i] == False and firstBool == True:
            continue
        elif check[i] == True:
            if firstBool == True:
                calQuan += int(quantity[i])
                firstBool = False
            else:
                inputQuan.append(calQuan)  
                inputValue.append(data["Value"][i])
                inputVariation.append(data["Variation"][i])
                calQuan = 0
                calQuan += int(quantity[i])
        elif data["Time"][i] == "09:01":            
            calQuan += int(quantity[i])
            inputTime.append("09:05:00")
            inputQuan.append(calQuan)
            inputValue.append(data["Value"][i - 4])
            inputVariation.append(data["Variation"][i - 4])
            calQuan = 0
        elif data["Time"][i] == "09:00":
            calQuan += int(quantity[i])
            inputTime.append("09:00:00")
            inputQuan.append(calQuan)
            inputValue.append(data["Value"][i])
            inputVariation.append(data["Variation"][i])
            calQuan = 0
        else:
            calQuan += int(quantity[i])
         
    inputTime.reverse()
    inputQuan.reverse()
    inputValue.reverse()
    inputVariation.reverse()
#    print(inputTime, inputQuan)
            
    
    
    
    #from openpyxl import load_workbook
    try:
        wb = xl.load_workbook(f'預估成交量-阿信_{date}.xlsx')
    except:
        wb = xl.load_workbook("預估成交量-阿信.xlsx")
    try:
        ws = wb[stockNo[no]]
    except:        
        ws = wb.copy_worksheet(wb["計算表"])
        ws.cell(row = 1, column = 13).value = stockNo[no]
        ws.cell(row = 1, column = 15).value = stockName[no]
        ws.title = str(stockNo[no])
    
    
    
    inputCount = 0
    for i in range(1,60):
        if str(ws.cell(row = i, column = 1).value) in inputTime:
            ws.cell(row = i, column = 3).value = inputQuan[inputCount]
            ws.cell(row = i, column = 6).value = inputValue[inputCount]
            ws.cell(row = i, column = 7).value = inputVariation[inputCount]
            inputCount += 1
    wb.save(f'預估成交量-阿信_{date}.xlsx')
    
    
import tkinter as tk
import os
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') 

pathDesktop = f'{desktop}\\預估量_{date}.xlsx'
try:
    wb.save(pathDesktop)
    print(f'{date}-{hour}:{minute}_更新完成')
except:
    window = tk.Tk()
    window.title("Stock")
    window.geometry("300x100+250+250")
    label = tk.Label(window, text = "Excel檔案開著，桌面檔案無法更新")
    label.pack()
    window.mainloop()

# 跳出視窗提示目前價格，但無法自動關閉視窗，這樣工作排程無法執行
# =============================================================================
# show = False
# checkMinu = int(minute)
# if checkMinu >= 0 and checkMinu < 10:
#     show = True
# elif checkMinu >= 15 and checkMinu < 20:
#     show = True
# elif checkMinu >= 30 and checkMinu < 35:
#     show = True
# elif checkMinu >= 45 and checkMinu < 50:
#     show = True
# 
# if show == True:
#     outLbl = f'{date}-{nowTime} \n'
#     for i in range(len(stockNo)):
#         outLbl += f'{stockNo[i]}, {stockName[i]}, {stockValue[i]} \n'
#     window = tk.Tk()
#     window.title("Stock")
#     window.geometry("300x300+250+250")
#     label = tk.Label(window, text = outLbl)
#     label.pack()
#     window.mainloop()
# =============================================================================

# =============================================================================
# print(stockNo)
# print(stockName)
# print(stockValue)
# =============================================================================


# =============================================================================
# b8 = ws.cell(row=8, column=2)
# #b1 = ws["B1"]
# print(b8.value)
# =============================================================================




# =============================================================================
# titles = root.find_all("div", class_ = "title") # 尋找 class = "title" 的 div 標籤
# for title in titles:
#     if title.a != None: # 如果標題包含 a 標籤 (沒有被刪除) print出來
#         print(title.a.string)
# =============================================================================
